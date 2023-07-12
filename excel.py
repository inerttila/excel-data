import os
import openpyxl
from tkinter import Tk, simpledialog, Label, StringVar, Button, OptionMenu, Entry, Text
from openpyxl.styles import Font, PatternFill
from tkcalendar import Calendar, DateEntry
import tkinter.messagebox as messagebox

# Set the file path and name
file_path = '/Users/skaitech/Desktop/EXCEL/your_file.xlsx'

# Extract the directory path from the file path
directory = os.path.dirname(file_path)

# Create the directory if it doesn't exist
os.makedirs(directory, exist_ok=True)

# Load the workbook or create a new one if it doesn't exist
try:
    workbook = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# Select the worksheet (replace 'Sheet1' with the actual sheet name)
worksheet = workbook['Sheet']

# Define the category labels
categories = ['Date', 'Service Line', 'Type of Service', 'Company', 'Task', 'Hours', 'Notes']

# Apply font and fill styles to the header row
header_font = Font(bold=True, color="ffffff")  # Bold font with white color
header_fill = PatternFill(start_color="061c43", end_color="061c43", fill_type="solid")  # Black background color

# Add category labels and apply styles to the header row
for col_num, category in enumerate(categories, start=1):
    cell = worksheet.cell(row=1, column=col_num, value=category)
    cell.font = header_font
    cell.fill = header_fill

# Define the predefined company options
company_options = ['Skaitech', '3DSkai']

# Define the predefined service line options
service_options = ['Develop', 'Drones/Robotics', 'IoT/Automation', 'Security/AI', 'Immersive Technologies', 'Training', 'Research']

# Define the predefined type of service options
type_of_service_options = ['Software', 'Hardware', 'Other']

# Define the predefined task options
task_options = ['Development', 'Control', 'Research', 'Testing', 'Other']

# Dictionary to store the input values
input_values = {}

# Function to handle the selection process for each category
def select_option(category, options):
    try:
        value = simpledialog.askstring("Data Input", f"Enter {category}:")
        if value in options:
            input_values[category].set(value)
        else:
            raise ValueError("Invalid option selected.")
    except ValueError as e:
        messagebox.showerror("Error", str(e))

# Function to handle the calendar selection
def select_date():
    def on_date_selected():
        selected_date = cal.selection_get()
        input_values['Date'].set(selected_date.strftime("%Y-%m-%d"))
        date_label.config(text=input_values['Date'].get())
        top.destroy()

    top = Tk()
    top.title("Select Date")

    # Define the position of the calendar within the pop-up window
    calendar_position_row = 0
    calendar_position_column = 1

    cal = Calendar(top)
    cal.grid(row=calendar_position_row, column=calendar_position_column, padx=10, pady=10)

    confirm_button = Button(top, text="Confirm", command=on_date_selected)
    confirm_button.grid(row=calendar_position_row + 1, column=calendar_position_column, padx=10, pady=10)

    top.mainloop()

# Function to handle validation for the Hours field
def validate_hours_input(value):
    if value and not value.replace('.', '').isdigit():
        return False
    return True

# Function to handle validation for the Notes field
def validate_notes_input(value):
    # Add your custom validation logic for the Notes field here
    return True

# Function to handle the confirmation button click
def confirm_input():
    try:
        for i, category in enumerate(categories):
            value = input_values[category].get()
            if category == 'Date' and not value:
                raise ValueError("Date must be selected.")
            if category == 'Hours' and not validate_hours_input(value):
                raise ValueError("Invalid Hours input.")
            if category == 'Notes' and not validate_notes_input(value):
                raise ValueError("Invalid Notes input.")
            worksheet.cell(row=2, column=i + 1, value=value)
        workbook.save(file_path)
        window.destroy()
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Create the main window
window = Tk()
window.title("Data Input")

# Configure pop-up window dimensions
window_width = 500
window_height = 300
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
x_coordinate = (screen_width - window_width) // 2
y_coordinate = (screen_height - window_height) // 2
window.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")


for i, category in enumerate(categories):
    # Create label
    label = Label(window, text=category)
    label.grid(row=i, column=0)

    # Create option menu for categories with predefined options
    if category == 'Company':
        input_values[category] = StringVar(window, company_options[0])
        option_menu = OptionMenu(window, input_values[category], *company_options)
        option_menu.grid(row=i, column=1)  # Display option menu for Task category
    
    elif category == 'Service Line':
        input_values[category] = StringVar(window, service_options[0])
        option_menu = OptionMenu(window, input_values[category], *service_options)
        option_menu.grid(row=i, column=1)  # Display option menu for Task category
    
    elif category == 'Type of Service':
        input_values[category] = StringVar(window, type_of_service_options[0])
        option_menu = OptionMenu(window, input_values[category], *type_of_service_options)
        option_menu.grid(row=i, column=1)  # Display option menu for Task category
        
    elif category == 'Task':
        input_values[category] = StringVar(window, task_options[0])
        option_menu = OptionMenu(window, input_values[category], *task_options)
        option_menu.grid(row=i, column=1)  # Display option menu for Task category
    elif category == 'Date':
        input_values[category] = StringVar(window)
        select_date_button = Button(window, text="Select Date", command=select_date)
        select_date_button.grid(row=i, column=1)
        date_label = Label(window, textvariable=input_values[category])
        date_label.grid(row=i, column=2)
    elif category == 'Hours':
        input_values[category] = StringVar(window)
        vcmd = (window.register(validate_hours_input), '%P')
        entry = Entry(window, textvariable=input_values[category], validate="key", validatecommand=vcmd)
        entry.grid(row=i, column=1)
    elif category == 'Notes':
        input_values[category] = StringVar(window)
        vcmd = (window.register(validate_notes_input), '%P')
        entry = Text(window, height=4, width=30)
        entry.grid(row=i, column=1, columnspan=2)
    else:
        input_values[category] = StringVar(window)
        entry = Entry(window, textvariable=input_values[category])
        entry.grid(row=i, column=1)

# Add confirm button
confirm_button = Button(window, text="Confirm", command=confirm_input)
confirm_button.grid(row=len(categories), column=0, columnspan=2, pady=10)

window.mainloop()
